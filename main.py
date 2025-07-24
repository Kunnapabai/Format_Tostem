#!/usr/bin/env python3
"""
Excel Color Extractor - Command Line Interface
รับ arguments จาก server.py และคืนผลลัพธ์เป็น JSON
"""

import os
import re
import math
import uuid
import shutil
import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
# Ensure pandas and openpyxl are installed

class ColorExtractor:
    def __init__(self, job_id: str):
        self.job_id = job_id
        
    def to_number(self, val):
        """Convert value to number, removing commas"""
        try:
            if val is None:
                return None
            
            str_val = str(val).strip()
            # Remove comma, space, and special characters
            clean_val = re.sub(r'[,\s]', '', str_val)
            clean_val = re.sub(r'[^\d.-]', '', clean_val)
            
            if not clean_val or clean_val in ['', '-', '.']:
                return None
                
            f = float(clean_val)
            if math.isnan(f):
                return None
            return int(f) if f.is_integer() else f
        except:
            return None

    def normalize_rgb(self, fill):
        """Convert ARGB color to RGB hex format - แก้ไขให้อ่านสีที่ถูกต้อง"""
        if not fill:
            return "FFFFFF"
        
        # ตรวจสอบ patternType ก่อน - เฉพาะ solid fill เท่านั้น
        if hasattr(fill, 'patternType') and fill.patternType:
            pattern_value = fill.patternType.value if hasattr(fill.patternType, 'value') else str(fill.patternType)
            # ถ้าไม่ใช่ solid pattern ให้ถือว่าไม่มีสี
            if pattern_value != 'solid':
                return "FFFFFF"
        else:
            # ถ้าไม่มี patternType ให้ถือว่าไม่มีสี
            return "FFFFFF"
        
        # รายการสีที่ไม่ต้องการ (Excel theme colors) - ไม่รวม 92CDDC
        excluded_colors = [
            "00000000",  # สีใส
            "F2F2F2"
        ]
        
        color_found = ""
        
        # Check fgColor
        if hasattr(fill, 'fgColor') and fill.fgColor:
            if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                color_str = str(fill.fgColor.rgb).upper()
                if color_str == "00000000":
                    return "FFFFFF"
                elif len(color_str) == 8:
                    color_found = color_str[2:]
                elif len(color_str) == 6:
                    color_found = color_str
        
        # Check bgColor
        if not color_found and hasattr(fill, 'bgColor') and fill.bgColor:
            if hasattr(fill.bgColor, 'rgb') and fill.bgColor.rgb:
                color_str = str(fill.bgColor.rgb).upper()
                if color_str == "00000000":
                    return "FFFFFF"
                elif len(color_str) == 8:
                    color_found = color_str[2:]
                elif len(color_str) == 6:
                    color_found = color_str
        
        # ตรวจสอบว่าเป็นสีที่ไม่ต้องการหรือไม่
        if color_found in excluded_colors:
            return "FFFFFF"
        
        return color_found if color_found else "FFFFFF"

    def find_thickness_matrix_in_column_a(self, ws, raw, thickness_num):
        """Find matrix with specific thickness label - หาจากคอลัมน์ A เท่านั้น"""
        thickness_patterns = [
            rf"Thk\.{thickness_num}",
            rf"\b{thickness_num}\b",
            rf"Thickness\s*{thickness_num}",
            rf"หนา\s*{thickness_num}",
            rf"ชั้น\s*{thickness_num}",
            rf"ระดับ\s*{thickness_num}"
        ]
        
        # หา thickness header ในคอลัมน์ A เท่านั้น (column index 0)
        for r in range(raw.shape[0]):
            if raw.shape[1] > 0:  # ตรวจสอบว่ามีคอลัมน์ A
                cell_val = str(raw.iat[r, 0]).strip() if raw.iat[r, 0] is not None else ""
                for pattern in thickness_patterns:
                    if re.search(pattern, cell_val, re.IGNORECASE):
                        print(f"   ✅ พบ {thickness_num} matrix ที่ row={r+1}, col=A (คอลัมน์ A)")
                        return r
        
        return None

    def find_main_matrix(self, ws, raw):
        """Find main matrix (1 or h/w header) - หา 1 จากคอลัมน์ A, h/w จากทั่วไป"""
        # หาจาก 1 header ในคอลัมน์ A เท่านั้น
        for r in range(raw.shape[0]):
            if raw.shape[1] > 0:  # ตรวจสอบว่ามีคอลัมน์ A
                cell_val = str(raw.iat[r, 0]).strip() if raw.iat[r, 0] is not None else ""
                # หา 1 header ในคอลัมน์ A
                if re.search(r"\b1\b", cell_val, re.IGNORECASE):
                    print(f"   ✅ พบ 1 matrix (main) ที่ row={r+1}, col=A (คอลัมน์ A)")
                    return r, 0  # ส่งคืน column = 0 (คอลัมน์ A)
        
        # ถ้าไม่พบ 1 header ให้หา h/w header แทน (ค้นหาทั่วไป - backward compatibility)
        for r in range(raw.shape[0]):
            for c in range(raw.shape[1]):
                if raw.iat[r, c] is None:
                    continue
                if isinstance(raw.iat[r, c], str):
                    if re.search(r"\bh\s*/\s*w\b", raw.iat[r, c], re.IGNORECASE):
                        print(f"   ✅ พบ h/w matrix (fallback) ที่ row={r+1}, col={c+1}")
                        return r, c
        
        return None, None

    def read_color_matrix_with_thickness_row(self, ws, raw, hr_main, hc_main, hr_thick, widths, heights, matrix_name=""):
        """อ่านสีจาก thickness row โดยใช้ position ของ main matrix"""
        print(f"     🔍 {matrix_name}: อ่านสีจาก thickness row {hr_thick+1}")
        print(f"     📍 Main matrix: row={hr_main+1}, col={hc_main+1}")
        print(f"     📍 Thickness header: row={hr_thick+1}, col=A")
        
        colors = {}
        
        # ลอง offset หลายแบบเหมือนฟังก์ชัน auto-offset เดิม
        best_colors = {}
        max_valid_colors = 0
        best_offset = (1, 1)
        
        # ลอง offset ต่างๆ โดยเริ่มจาก thickness row
        for row_offset in [1, 2, 3]:
            for col_offset in [1, 2, 3]:
                test_colors = {}
                valid_count = 0
                
                # ทดสอบเฉพาะ 4 เซลล์แรก
                for i_h in range(min(2, len(heights))):
                    for i_w in range(min(2, len(widths))):
                        h, w = heights[i_h], widths[i_w]
                        try:
                            # เริ่มจาก thickness row + offset
                            excel_row = hr_thick + row_offset + i_h
                            excel_col = hc_main + col_offset + i_w  # ใช้ col ของ main matrix
                            
                            if excel_row <= ws.max_row and excel_col <= ws.max_column:
                                cell = ws.cell(row=excel_row, column=excel_col)
                                color = self.normalize_rgb(cell.fill)
                                test_colors[(h, w)] = color
                                if color != "FFFFFF":
                                    valid_count += 1
                            else:
                                test_colors[(h, w)] = "FFFFFF"
                        except Exception as e:
                            test_colors[(h, w)] = "FFFFFF"
                
                # ถ้า offset นี้ให้ผลดีกว่า
                if valid_count > max_valid_colors:
                    max_valid_colors = valid_count
                    best_offset = (row_offset, col_offset)
                    print(f"       🎯 offset +{row_offset},+{col_offset}: {valid_count} สี")
        
        # ใช้ offset ที่ดีที่สุดเพื่ออ่านทั้ง matrix
        row_offset, col_offset = best_offset
        print(f"     ✅ ใช้ offset สำหรับ {matrix_name}: +{row_offset},+{col_offset}")
        
        for i_h, h in enumerate(heights):
            for i_w, w in enumerate(widths):
                try:
                    excel_row = hr_thick + row_offset + i_h
                    excel_col = hc_main + col_offset + i_w
                    
                    if excel_row <= ws.max_row and excel_col <= ws.max_column:
                        cell = ws.cell(row=excel_row, column=excel_col)
                        color = self.normalize_rgb(cell.fill)
                        best_colors[(h, w)] = color
                    else:
                        best_colors[(h, w)] = "FFFFFF"
                except:
                    best_colors[(h, w)] = "FFFFFF"
        
        # แสดงผลสรุป
        colored_count = sum(1 for color in best_colors.values() if color != "FFFFFF")
        print(f"     📊 {matrix_name}: อ่านได้ {colored_count}/{len(best_colors)} เซลล์ที่มีสี")
        
        return best_colors

    def read_color_matrix(self, ws, raw, hr, hc, widths, heights):
        """Read colors from matrix - ใช้ offset มาตรฐาน"""
        color_map = {}
        
        for i_h, h in enumerate(heights):
            for i_w, w in enumerate(widths):
                try:
                    excel_row = hr + 2 + i_h
                    excel_col = hc + 2 + i_w
                    
                    cell = ws.cell(row=excel_row, column=excel_col)
                    color = self.normalize_rgb(cell.fill)
                    color_map[(h, w)] = color
                except Exception:
                    color_map[(h, w)] = "FFFFFF"
                    continue
        
        return color_map

    def scan_all_matrices_in_file(self, xls, wb):
        """สแกนทุกชีตเพื่อหาจำนวน matrix สูงสุด"""
        max_matrices = 1  # อย่างน้อยต้องมี matrix 1
        max_sheet = ""
        all_sheet_matrices = {}
        
        print("\n🔍 สแกนทุกชีตเพื่อหาจำนวน matrix...")
        
        for sheet_name in xls.sheet_names:
            if sheet_name.strip().lower() == "สารบัญ":
                continue
                
            print(f"   📋 สแกน Sheet: {sheet_name}")
            
            try:
                raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine="openpyxl")
                ws = wb[sheet_name]
                
                # หา main matrix
                hr, hc = self.find_main_matrix(ws, raw)
                if hr is None:
                    print(f"      ❌ ไม่พบ main matrix ใน {sheet_name}")
                    all_sheet_matrices[sheet_name] = []
                    continue
                
                # หา matrices ทั้งหมดในชีตนี้
                found_matrices = [1]  # 1 เป็น main matrix เสมอ
                
                for thickness in range(2, 20):  # ตรวจหาสูงสุด 20 matrices
                    hr_thick = self.find_thickness_matrix_in_column_a(ws, raw, thickness)
                    if hr_thick is not None:
                        found_matrices.append(thickness)
                        print(f"      ✅ พบ matrix {thickness}")
                    else:
                        # ถ้าไม่เจอ matrix ลำดับถัดไป ให้หยุดค้นหา
                        break
                
                all_sheet_matrices[sheet_name] = found_matrices
                matrix_count = len(found_matrices)
                print(f"      📊 รวม {matrix_count} matrices: {found_matrices}")
                
                # อัพเดทจำนวน matrix สูงสุด
                if matrix_count > max_matrices:
                    max_matrices = matrix_count
                    max_sheet = sheet_name
                    print(f"      🏆 ชีต {sheet_name} มี matrix เยอะที่สุด: {matrix_count} matrices")
                    
            except Exception as e:
                print(f"      ❌ Error สแกน {sheet_name}: {e}")
                all_sheet_matrices[sheet_name] = []
        
        print(f"\n🎯 ผลการสแกน:")
        print(f"   🏆 ชีตที่มี matrix เยอะที่สุด: {max_sheet} ({max_matrices} matrices)")
        print(f"   📋 รายละเอียดทุกชีต:")
        for sheet, matrices in all_sheet_matrices.items():
            if matrices:
                print(f"      - {sheet}: {len(matrices)} matrices {matrices}")
            else:
                print(f"      - {sheet}: ไม่พบ matrix")
        
        return max_matrices, all_sheet_matrices

    def process_file(self, input_file: str, output_dir: str, original_filename: str = None):
        """Process the Excel file"""
        try:
            if original_filename:
                base_name = os.path.splitext(original_filename)[0]
            else:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                # ลบ UUID ออกจากชื่อไฟล์ (UUID format: 8-4-4-4-12 characters)
                uuid_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_'
                base_name = re.sub(uuid_pattern, '', base_name)
            
            xls = pd.ExcelFile(input_file, engine="openpyxl")
            wb = load_workbook(input_file, data_only=True)
            
            # สแกนทุกชีตเพื่อหาจำนวน matrix สูงสุด
            max_matrices_count, all_sheet_matrices = self.scan_all_matrices_in_file(xls, wb)
            
            # สร้าง template คอลัมน์ตามจำนวน matrix สูงสุด
            matrix_columns = []
            for i in range(1, max_matrices_count + 1):
                matrix_columns.append(f"{i}_Color")
            
            print(f"\n📝 จะสร้างคอลัมน์: {matrix_columns}")
            
            price_rows = []
            type_rows = []
            price_id = 1
            type_id = 1
            
            # Track processing results
            processed_sheets = 0
            skipped_sheets = []
            warnings = []
            
            for sheet in xls.sheet_names:
                # ตรวจสอบ Sheet สารบัญ
                if sheet.strip().lower() == "สารบัญ":
                    skipped_sheets.append({"sheet": sheet, "reason": "ข้าม Sheet สารบัญ"})
                    print(f"   ⚠️ ข้าม Sheet: {sheet} (สารบัญ)")
                    continue
                
                print(f"\n🔍 ประมวลผล Sheet: {sheet}")
                
                # ใช้ข้อมูลจากการสแกน
                available_matrices = all_sheet_matrices.get(sheet, [])
                if not available_matrices:
                    error_msg = "ไม่พบ matrix ใดๆ"
                    print(f"   ❌ {error_msg} ใน {sheet}")
                    skipped_sheets.append({"sheet": sheet, "reason": error_msg})
                    continue
                
                raw = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
                ws = wb[sheet]
                
                # Find Glass_QTY and Description
                sheet_glass_qty = 1
                sheet_description = ""
                
                for r in range(raw.shape[0]):
                    for c in range(raw.shape[1] - 1):
                        if raw.iat[r, c] is None:
                            continue
                        cell = str(raw.iat[r, c]).strip()
                        low = cell.lower()
                        
                        if low in ("glass_qty", "glass qty"):
                            next_cell = raw.iat[r, c + 1]
                            qty = self.to_number(next_cell)
                            if qty is not None:
                                sheet_glass_qty = qty
                            
                        elif low == "description":
                            desc = raw.iat[r, c + 1]
                            if desc is not None:
                                sheet_description = str(desc).strip()
                
                # Find main matrix (1 or h/w header)
                hr, hc = self.find_main_matrix(ws, raw)
                
                if hr is None or hc is None:
                    error_msg = "ไม่พบ main matrix"
                    print(f"   ❌ {error_msg} ใน {sheet}")
                    skipped_sheets.append({"sheet": sheet, "reason": error_msg})
                    continue
                
                # Read widths and heights from main matrix
                widths = []
                for c in range(hc + 1, raw.shape[1]):
                    v = self.to_number(raw.iat[hr, c])
                    if v is None:
                        break
                    widths.append(v)
                
                heights = []
                for r in range(hr + 1, raw.shape[0]):
                    h_val = self.to_number(raw.iat[r, hc])
                    if h_val is None:
                        break
                    heights.append(h_val)
                
                if not widths or not heights:
                    error_msg = "ไม่พบ dimensions (ความกว้าง/ความสูง)"
                    print(f"   ❌ {error_msg} ใน {sheet}")
                    skipped_sheets.append({"sheet": sheet, "reason": error_msg})
                    continue
                
                print(f"   📊 Dimensions: {len(heights)} heights x {len(widths)} widths")
                print(f"   🎯 Matrices ในชีตนี้: {available_matrices}")
                
                # อ่านสีจาก matrices ที่มี
                matrix_colors = {}
                
                # อ่าน matrix 1 (main matrix)
                if 1 in available_matrices:
                    matrix_colors[1] = self.read_color_matrix(ws, raw, hr, hc, widths, heights)
                    print(f"   🎨 1 (main matrix): {len(matrix_colors[1])} colors")
                
                # อ่าน matrices อื่นๆ
                for thickness in available_matrices:
                    if thickness == 1:
                        continue  # ข้าม matrix 1 เพราะอ่านไปแล้ว
                    
                    hr_thick = self.find_thickness_matrix_in_column_a(ws, raw, thickness)
                    if hr_thick is not None:
                        colors = self.read_color_matrix_with_thickness_row(
                            ws, raw, hr, hc, hr_thick, widths, heights, f"{thickness}"
                        )
                        matrix_colors[thickness] = colors
                        print(f"   🎨 {thickness}: {len(colors)} colors อ่านได้")
                
                # Create Type record
                type_rows.append({
                    "ID": type_id,
                    "Serie": base_name,
                    "Type": sheet.strip(),
                    "Description": sheet_description,
                    "width_min": min(widths),
                    "width_max": max(widths),
                    "height_min": min(heights),
                    "height_max": max(heights),
                })
                type_id += 1
                
                # Create Price records with consistent columns
                sheet_price_count = 0
                for i_h, h in enumerate(heights):
                    for i_w, w in enumerate(widths):
                        # อ่านราคาจาก main matrix (1)
                        raw_price = raw.iat[hr + 1 + i_h, hc + 1 + i_w]
                        p = self.to_number(raw_price)
                        if p is None:
                            continue
                        
                        # สร้าง price record พร้อมคอลัมน์ตามมาตรฐาน
                        price_record = {
                            "ID": price_id,
                            "Serie": base_name,
                            "Type": sheet.strip(),
                            "Width": w,
                            "Height": h,
                            "Price": p,
                            "Glass_QTY": sheet_glass_qty,
                        }
                        
                        # เพิ่มคอลัมน์สีทุกคอลัมน์ตามมาตรฐาน (เติม FFFFFF ถ้าไม่มี)
                        for i in range(1, max_matrices_count + 1):
                            color_key = f"{i}_Color"
                            if i in matrix_colors:
                                color_value = matrix_colors[i].get((h, w), "FFFFFF")
                            else:
                                color_value = "FFFFFF"  # ไม่มี matrix นี้ในชีตนี้
                            price_record[color_key] = color_value
                        
                        price_rows.append(price_record)
                        price_id += 1
                        sheet_price_count += 1
                
                processed_sheets += 1
                print(f"   ✅ สร้าง {sheet_price_count} price records สำหรับ {sheet}")
            
            # Ensure output directory exists
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)
            
            # Save output files
            price_file = output_path / f"Price_{self.job_id}.xlsx"
            type_file = output_path / f"Type_{self.job_id}.xlsx"
            
            pd.DataFrame(price_rows).to_excel(price_file, index=False)
            pd.DataFrame(type_rows).to_excel(type_file, index=False)
            
            print(f"\n✅ เสร็จสิ้น: {len(price_rows)} price records, {len(type_rows)} type records")
            print(f"📋 คอลัมน์ที่สร้าง: {matrix_columns}")
            
            return {
                "price_file": str(price_file),
                "type_file": str(type_file),
                "total_records": len(price_rows),
                "processed_sheets": processed_sheets,
                "skipped_sheets": skipped_sheets,
                "warnings": warnings
            }
            
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            raise Exception(f"Processing failed: {str(e)}")

def main():
    """Main function to handle command line arguments"""
    parser = argparse.ArgumentParser(description='Excel Color Extractor - Matrix Mode')
    parser.add_argument('--input', required=True, help='Input Excel file path')
    parser.add_argument('--job-id', required=True, help='Job ID for output files')
    parser.add_argument('--output-dir', default='outputs', help='Output directory')
    parser.add_argument('--original-filename', help='Original filename for base name extraction')
    
    args = parser.parse_args()
    
    try:
        # Validate input file
        if not os.path.exists(args.input):
            raise FileNotFoundError(f"Input file not found: {args.input}")
        
        if not args.input.lower().endswith('.xlsx'):
            raise ValueError("Input file must be an .xlsx file")
        
        print(f"🚀 Starting Excel Color Extractor...")
        print(f"📄 Input file: {args.input}")
        print(f"🆔 Job ID: {args.job_id}")
        print(f"📁 Output directory: {args.output_dir}")
        if args.original_filename:
            print(f"📝 Original filename: {args.original_filename}")
        
        # Process the file
        extractor = ColorExtractor(args.job_id)
        result = extractor.process_file(
            input_file=args.input,
            output_dir=args.output_dir,
            original_filename=args.original_filename
        )
        
        # Output result as JSON for server.py to parse
        print(json.dumps(result))
        
    except Exception as e:
        print(f"❌ Error: {str(e)}", file=sys.stderr)
        exit(1)

if __name__ == "__main__":
    import sys
    main()