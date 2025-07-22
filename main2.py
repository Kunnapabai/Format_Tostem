import pandas as pd
import os
import re
from typing import List, Dict, Tuple, Optional
import logging
from openpyxl import load_workbook
from flask import Flask, request, jsonify, send_file, render_template_string
import uuid
import time
import shutil
from werkzeug.utils import secure_filename



# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self, input_file: str, original_filename: str = None):
        self.input_file = input_file
        self.original_filename = original_filename
        self.price_records: List[Dict] = []
        self.type_records: List[Dict] = []
        self.price_id = 1
        self.type_id = 1
        self.description_map: Dict[str, str] = {}
        
        # Extract series name from filename
        self.series_name = self.extract_series_from_filename()
        print(f"📱 ชื่อ Serie: {self.series_name}")
        
        # Cache for optimized reading
        self._wb = None
        self._sheets_cache = {}
    
    def extract_series_from_filename(self) -> str:
        """ดึงชื่อ series จากชื่อไฟล์ โดยจัดการกับ UUID และ timestamp"""
        if self.original_filename:
            # ใช้ชื่อไฟล์ต้นฉบับ
            base_name = os.path.splitext(self.original_filename)[0]
        else:
            # ใช้ชื่อไฟล์ปัจจุบัน
            base_name = os.path.splitext(os.path.basename(self.input_file))[0]
        
        # ลบ timestamp pattern (YYYYMMDD_HHMMSS_)
        timestamp_pattern = r'^\d{8}_\d{6}_[a-f0-9]{8}_'
        base_name = re.sub(timestamp_pattern, '', base_name)
        
        # ลบ UUID pattern (8-4-4-4-12 characters)
        uuid_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_'
        base_name = re.sub(uuid_pattern, '', base_name)
        
        # ลบ job_id pattern ที่อาจมี
        job_id_pattern = r'^[a-f0-9]{8}_'
        base_name = re.sub(job_id_pattern, '', base_name)
        
        # ลบ prefix/suffix ที่ไม่ต้องการ
        suffixes_to_remove = ['_data', '_price', '_export', '_backup', '_processed']
        prefixes_to_remove = ['data_', 'price_', 'export_', 'backup_', 'processed_']
        
        # ลบ suffix
        for suffix in suffixes_to_remove:
            if base_name.lower().endswith(suffix):
                base_name = base_name[:-len(suffix)]
                break
        
        # ลบ prefix
        for prefix in prefixes_to_remove:
            if base_name.lower().startswith(prefix):
                base_name = base_name[len(prefix):]
                break
        
        # ลบช่องว่างและอักขระพิเศษ
        base_name = base_name.strip().replace(' ', '_')
        
        return base_name
    
    def validate_file(self) -> bool:
        """Validate that the input file exists and is accessible"""
        if not os.path.exists(self.input_file):
            logger.error(f"ไม่เจอไฟล์ {self.input_file}")
            return False
        return True
    
    def get_optimized_workbook(self):
        """Get cached workbook with optimized settings"""
        if self._wb is None:
            logger.info("Loading workbook with optimized settings...")
            self._wb = load_workbook(
                self.input_file, 
                read_only=True,  # Much faster
                data_only=True,  # Get calculated values
                keep_links=False  # Don't load external links
            )
        return self._wb
    
    def read_sheet_optimized(self, sheet_name_or_index, **kwargs):
        """Read sheet with optimized pandas settings"""
        cache_key = f"{sheet_name_or_index}_{str(kwargs)}"
        if cache_key not in self._sheets_cache:
            logger.info(f"Loading sheet: {sheet_name_or_index}")
            
            # Use pandas default engine (openpyxl) without conflicting parameters
            self._sheets_cache[cache_key] = pd.read_excel(
                self.input_file,
                sheet_name=sheet_name_or_index,
                engine='openpyxl',
                **kwargs
            )
        return self._sheets_cache[cache_key]
    
    def load_descriptions_from_sheet2(self) -> bool:
        """Load descriptions from sheet2 mapping Type to Description - OPTIMIZED"""
        try:
            print("📖 กำลังอ่าน sheet2 สำหรับ descriptions...")
            logger.info("Loading descriptions from sheet2...")
            # Use optimized reading
            df_sheet2 = self.read_sheet_optimized(1, dtype=str)
            
            # Strip whitespace from column names
            df_sheet2.columns = df_sheet2.columns.str.strip()
            
            # Find Type and Description columns
            type_col = None
            desc_col = None
            
            for i, col in enumerate(df_sheet2.columns):
                if 'type' in str(col).lower():
                    type_col = col
                    if i + 1 < len(df_sheet2.columns):
                        desc_col = df_sheet2.columns[i + 1]
                    break
            
            if type_col is None or desc_col is None:
                print("⚠️ ไม่พบคอลัมน์ Type หรือ Description ใน sheet2")
                logger.warning("ไม่พบคอลัมน์ Type หรือ Description ใน sheet2")
                return False
            
            print(f"✅ พบคอลัมน์: Type='{type_col}', Description='{desc_col}'")
            
            # Create mapping efficiently using vectorized operations
            valid_mask = (df_sheet2[type_col].notna()) & (df_sheet2[type_col] != 'nan')
            valid_data = df_sheet2[valid_mask]
            
            for _, row in valid_data.iterrows():
                type_name = str(row[type_col]).strip()
                desc_text = str(row[desc_col]).strip() if pd.notna(row[desc_col]) else ''
                if type_name:
                    self.description_map[type_name] = desc_text
            
            print(f"✅ โหลด descriptions สำเร็จ: {len(self.description_map)} รายการ")
            logger.info(f"Loaded {len(self.description_map)} descriptions from sheet2")
            return True
            
        except Exception as e:
            print(f"❌ Error อ่าน sheet2: {e}")
            logger.error(f"Error loading sheet2: {e}")
            return False
    
    def update_type_descriptions(self):
        """Update type records with descriptions from sheet2"""
        for record in self.type_records:
            type_name = record['Type']
            description = self.description_map.get(type_name, '')
            record['Description'] = description
            
        logger.info("Updated Type descriptions")
    
    def clean_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize multi-level headers"""
        clean_cols = []
        for top, sub in df.columns:
            top_str = str(top).strip() if pd.notna(top) else ''
            sub_str = str(sub).strip() if pd.notna(sub) else ''
            clean_cols.append((top_str, sub_str))
        df.columns = pd.MultiIndex.from_tuples(clean_cols)
        return df
    
    def read_cell_background_color_optimized(self, sheet_name: str, row: int, col: int) -> str:
        """Read background color from Excel cell - OPTIMIZED"""
        try:
            wb = self.get_optimized_workbook()
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # Use 1-based indexing for openpyxl
            cell = ws.cell(row=row + 1, column=col + 1)
            
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                color = str(cell.fill.start_color.rgb)
                # Remove 'FF' prefix if present (alpha channel)
                if len(color) == 8 and color.startswith('FF'):
                    color = color[2:]
                
                # Check for empty colors
                if color == '00000000' or color == '000000' or not color:
                    return 'FFFFFF'
                
                return color
            return 'FFFFFF'
            
        except Exception as e:
            logger.warning(f"Cannot read cell color: {e}")
            return 'FFFFFF'
    
    def find_dimension_mode(self, sub_df: pd.DataFrame) -> Optional[str]:
        """Find the dimension mode (W first priority, then H)"""
        if 'W' in sub_df.columns:
            return 'W'
        elif 'H' in sub_df.columns:
            return 'H'
        return None
    
    def process_width_data(self, table_name: str, vals: pd.DataFrame, 
                          sheet_name: str = None) -> Tuple[float, float]:
        """Process width-based pricing data - OPTIMIZED"""
        w_vals = vals['W'].astype(float)
        p_vals = vals['Price'].astype(float)
        wmin, wmax = w_vals.min(), w_vals.max()
        
        # Pre-calculate color column index
        price_col_idx = list(vals.columns).index('Price')
        
        # Process in batch for better performance
        for idx, (w, p) in enumerate(zip(w_vals, p_vals)):
            original_idx = vals.index[idx]
            
            # Read color optimized
            color = 'FFFFFF'
            if sheet_name:
                color = self.read_cell_background_color_optimized(
                    sheet_name, original_idx + 2, price_col_idx
                )
            
            self.price_records.append({
                'ID': self.price_id,
                'Serie': self.series_name,  # เปลี่ยนจาก 'Series': 0
                'Type': table_name,
                'Width': w,
                'Height': 0,
                'Price': p,
                'Glass_QTY': 0,
                'Color': color
            })
            self.price_id += 1
        
        return wmin, wmax
    
    def process_height_data(self, table_name: str, vals: pd.DataFrame,
                           sheet_name: str = None) -> Tuple[float, float]:
        """Process height-based pricing data - OPTIMIZED"""
        h_vals = vals['H'].astype(float)
        p_vals = vals['Price'].astype(float)
        hmin, hmax = h_vals.min(), h_vals.max()
        
        # Pre-calculate color column index
        price_col_idx = list(vals.columns).index('Price')
        
        # Process in batch for better performance
        for idx, (h, p) in enumerate(zip(h_vals, p_vals)):
            original_idx = vals.index[idx]
            
            # Read color optimized
            color = 'FFFFFF'
            if sheet_name:
                color = self.read_cell_background_color_optimized(
                    sheet_name, original_idx + 2, price_col_idx
                )
            
            self.price_records.append({
                'ID': self.price_id,
                'Serie': self.series_name,  # เปลี่ยนจาก 'Series': 0
                'Type': table_name,
                'Width': 0,
                'Height': h,
                'Price': p,
                'Glass_QTY': 0,
                'Color': color
            })
            self.price_id += 1
        
        return hmin, hmax
    
    def add_type_record(self, table_name: str, wmin: float, wmax: float, 
                       hmin: float, hmax: float):
        """Add a type record with dimension ranges"""
        self.type_records.append({
            'ID': self.type_id,
            'Serie': self.series_name,  # เปลี่ยนจาก 'Series': 0
            'Type': table_name,
            'Description': '',  # Will be updated later
            'width_min': wmin,
            'width_max': wmax,
            'height_min': hmin,
            'height_max': hmax
        })
        self.type_id += 1
    
    def process_table(self, table_name: str, sub_df: pd.DataFrame, 
                     sheet_name: str = None) -> bool:
        """Process a single table from the Excel file - OPTIMIZED"""
        print(f"📊 ประมวลผล Table: {table_name}")
        
        # Clean column names
        sub_df.columns = sub_df.columns.str.strip()
        
        # Find dimension mode
        mode = self.find_dimension_mode(sub_df)
        if mode is None:
            print(f"⚠️ ข้าม {table_name}: ไม่มีคอลัมน์ W หรือ H")
            logger.warning(f"Skip {table_name}: no W or H column")
            return False
        
        print(f"✅ พบ dimension mode: {mode}")
        
        # Check for Price column
        if 'Price' not in sub_df.columns:
            print(f"⚠️ ข้าม {table_name}: ไม่มีคอลัมน์ Price")
            logger.warning(f"Skip {table_name}: no Price column")
            return False
        
        # Extract valid rows efficiently
        required_cols = [mode, 'Price']
        vals = sub_df[required_cols].dropna(how='any')
        
        if vals.empty:
            print(f"⚠️ ข้าม {table_name}: ไม่มีแถวข้อมูลครบ {mode} + Price")
            logger.warning(f"Skip {table_name}: no valid {mode} + Price rows")
            return False
        
        print(f"📋 พบข้อมูล {len(vals)} แถว")
        
        try:
            # Process based on mode
            if mode == 'W':
                print(f"🔄 ประมวลผล Width data สำหรับ {table_name}")
                wmin, wmax = self.process_width_data(table_name, vals, sheet_name)
                hmin = hmax = 0
                print(f"📏 Width range: {wmin} - {wmax}")
            else:  # mode == 'H'
                print(f"🔄 ประมวลผล Height data สำหรับ {table_name}")
                hmin, hmax = self.process_height_data(table_name, vals, sheet_name)
                wmin = wmax = 0
                print(f"📏 Height range: {hmin} - {hmax}")
            
            # Add type record
            self.add_type_record(table_name, wmin, wmax, hmin, hmax)
            print(f"✅ เสร็จสิ้น {table_name}: {len(vals)} แถว")
            logger.info(f"Processed {table_name}: {len(vals)} rows")
            return True
            
        except Exception as e:
            print(f"❌ Error ประมวลผล {table_name}: {e}")
            logger.error(f"Error processing {table_name}: {e}")
            return False
    
    def save_results(self, job_id: str) -> None:
        """Save processed data to Excel files with simple names"""
        if self.price_records:
            price_filename = 'Price.xlsx'
            pd.DataFrame(self.price_records).to_excel(price_filename, index=False)
            logger.info(f"Saved {len(self.price_records)} price records to {price_filename}")
        
        if self.type_records:
            type_filename = 'Type.xlsx'
            pd.DataFrame(self.type_records).to_excel(type_filename, index=False)
            logger.info(f"Saved {len(self.type_records)} type records to {type_filename}")
    
    def process(self, job_id: str) -> bool:
        """Main processing function - OPTIMIZED"""
        if not self.validate_file():
            return False
        
        try:
            print(f"🚀 เริ่มประมวลผล: {self.input_file}")
            logger.info(f"Starting optimized processing of {self.input_file}")
            
            # Get optimized workbook for color reading
            print("📂 กำลังเปิดไฟล์...")
            wb = self.get_optimized_workbook()
            sheet_name = wb.sheetnames[0]  # First sheet name
            print(f"✅ เปิดไฟล์สำเร็จ - Sheet หลัก: {sheet_name}")
            
            # Read main sheet with optimized settings
            print("📋 กำลังอ่าน main sheet...")
            logger.info("Loading main sheet...")
            df = self.read_sheet_optimized(0, header=[0, 1], dtype=str)
            print(f"✅ อ่าน main sheet สำเร็จ")
            
            # Clean headers
            print("🔧 กำลังทำความสะอาด headers...")
            df = self.clean_headers(df)
            
            # Filter out empty top-level columns
            df = df.loc[:, df.columns.get_level_values(0) != '']
            
            # Process each table in order
            print("🔄 เริ่มประมวลผลตารางต่างๆ...")
            processed_count = 0
            table_names = df.columns.get_level_values(0).unique()
            
            print(f"📊 พบ {len(table_names)} ตาราง: {list(table_names)}")
            
            for table_name in table_names:
                if self.process_table(table_name, df[table_name].copy(), sheet_name):
                    processed_count += 1
            
            print(f"✅ ประมวลผลตารางเสร็จสิ้น: {processed_count}/{len(table_names)}")
            
            # Load descriptions from sheet2
            print("📖 กำลังอ่าน descriptions จาก sheet2...")
            self.load_descriptions_from_sheet2()
            
            # Update type records with descriptions
            print("🔄 กำลังอัพเดท descriptions...")
            self.update_type_descriptions()
            
            # Save results with job_id
            print("💾 กำลังบันทึกผลลัพธ์...")
            self.save_results(job_id)
            
            print(f"🎉 ประมวลผลเสร็จสิ้น: {processed_count} ตาราง")
            print(f"📊 Price records: {len(self.price_records)}")
            print(f"📊 Type records: {len(self.type_records)}")
            
            logger.info(f"Processing complete: {processed_count} tables processed")
            return processed_count > 0
            
        except Exception as e:
            print(f"❌ Error: {e}")
            logger.error(f"Error during processing: {e}")
            return False
        finally:
            # Clean up resources
            if self._wb:
                self._wb.close()
                print("🔒 ปิดไฟล์แล้ว")

def process_multi_table_excel(input_file: str, job_id: str, original_filename: str = None) -> bool:
    """
    Process multi-table Excel file and generate Price.xlsx and Type.xlsx
    
    Args:
        input_file: Path to the input Excel file
        job_id: Unique job identifier for output files
        original_filename: Original filename before processing
        
    Returns:
        bool: True if processing was successful, False otherwise
    """
    processor = ExcelProcessor(input_file, original_filename)
    return processor.process(job_id)

# Flask Web Application
app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25MB
ALLOWED_EXTENSIONS = {'xlsx'}

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cleanup_old_files():
    """Clean up files older than 1 hour"""
    try:
        current_time = time.time()
        for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path):
                    if current_time - os.path.getctime(file_path) > 3600:  # 1 hour
                        os.remove(file_path)
                        logger.info(f"Cleaned up old file: {file_path}")
    except Exception as e:
        logger.error(f"Error during cleanup: {e}")

# Read the HTML template from index2.html
def load_html_template():
    try:
        with open('index2.html', 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        return """
        <html><body>
        <h1>Error: index2.html not found</h1>
        <p>Please make sure index2.html is in the same directory as main2.py</p>
        </body></html>
        """

@app.route('/')
def index():
    """Serve the main HTML page"""
    cleanup_old_files()
    html_template = load_html_template()
    return render_template_string(html_template)

@app.route('/api/process', methods=['POST'])
def process_file():
    """Process uploaded Excel file"""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'message': 'ไม่พบไฟล์'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'message': 'ไม่ได้เลือกไฟล์'}), 400
        
        # Validate file
        if not allowed_file(file.filename):
            return jsonify({'message': 'ประเภทไฟล์ไม่ถูกต้อง กรุณาอัพโหลดไฟล์ .xlsx'}), 400
        
        # Check file size
        file_content = file.read()
        if len(file_content) > MAX_FILE_SIZE:
            return jsonify({'message': 'ไฟล์ใหญ่เกินไป (สูงสุด 25MB)'}), 400
        file.seek(0)  # Reset file pointer
        
        # Generate job ID with timestamp for better naming
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_suffix = str(uuid.uuid4())[:8]  # Short UUID for uniqueness
        job_id = f"{timestamp}_{random_suffix}"
        
        # Save uploaded file
        original_filename = file.filename  # เก็บชื่อไฟล์ต้นฉบับ
        filename = secure_filename(file.filename)
        input_path = os.path.join(UPLOAD_FOLDER, f'{job_id}_{filename}')
        file.save(input_path)
        
        logger.info(f"Processing file: {filename} with job_id: {job_id}")
        
        # Record start time
        start_time = time.time()
        
        # Process the file with original filename
        success = process_multi_table_excel(input_path, job_id, original_filename)
        
        # Calculate processing time
        processing_time = time.time() - start_time
        
        # Clean up input file
        try:
            os.remove(input_path)
        except:
            pass
        
        if not success:
            return jsonify({
                'message': 'เกิดข้อผิดพลาดในการประมวลผล'
            }), 500
        
        # Count records in generated files
        price_count = 0
        type_count = 0
        
        try:
            price_file = 'Price.xlsx'
            type_file = 'Type.xlsx'
            
            if os.path.exists(price_file):
                price_count = len(pd.read_excel(price_file))
                # Copy to output folder with job_id for download tracking
                shutil.copy2(price_file, os.path.join(OUTPUT_FOLDER, f'Price_{job_id}.xlsx'))
                
            if os.path.exists(type_file):
                type_count = len(pd.read_excel(type_file))
                # Copy to output folder with job_id for download tracking
                shutil.copy2(type_file, os.path.join(OUTPUT_FOLDER, f'Type_{job_id}.xlsx'))
                
        except Exception as e:
            logger.error(f"Error moving files: {e}")
            return jsonify({'message': f'เกิดข้อผิดพลาดในการจัดการไฟล์: {str(e)}'}), 500
        
        logger.info(f"Processing completed successfully for job_id: {job_id}")
        
        return jsonify({
            'job_id': job_id,
            'total_records': price_count + type_count,
            'price_records': price_count,
            'type_records': type_count,
            'processed_sheets': 1,  # From main sheet
            'processing_time': processing_time,
            'message': 'ประมวลผลสำเร็จ'
        })
        
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return jsonify({'message': f'เกิดข้อผิดพลาดที่ไม่คาดคิด: {str(e)}'}), 500

@app.route('/api/download/<job_id>/<file_type>')
def download_file(job_id, file_type):
    """Download processed files"""
    try:
        if file_type == 'price':
            filename = f'Price_{job_id}.xlsx'
        elif file_type == 'type':
            filename = f'Type_{job_id}.xlsx'
        else:
            return jsonify({'message': 'ประเภทไฟล์ไม่ถูกต้อง'}), 400
        
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'message': 'ไม่พบไฟล์'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name='Price.xlsx' if file_type == 'price' else 'Type.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Download error: {e}")
        return jsonify({'message': f'เกิดข้อผิดพลาดในการดาวน์โหลด: {str(e)}'}), 500

@app.errorhandler(413)
def too_large(e):
    return jsonify({'message': 'ไฟล์ใหญ่เกินไป (สูงสุด 25MB)'}), 413

# Command line usage (original functionality)
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) == 3:
        # Command line mode
        print("🚀 เริ่มต้นโปรแกรม main2.py (Command Line Mode)")
        input_filename, job_id = sys.argv[1], sys.argv[2]
        print(f"📁 ไฟล์ Input: {input_filename}")
        print(f"🆔 Job ID: {job_id}")
        
        success = process_multi_table_excel(input_filename, job_id)
        
        if not success:
            print("❌ ERROR: processing failed")
            sys.exit(1)

        print("🎯 กำลังรวมผลลัพธ์...")
        # Print output in format expected by api.py
        print(f"MOVED_PRICE:Price.xlsx")
        print(f"MOVED_TYPE:Type.xlsx")
        
        # Count records in generated files
        try:
            price_count2 = len(pd.read_excel('Price.xlsx'))
            type_count2 = len(pd.read_excel('Type.xlsx'))
            print(f"PRICE_COUNT:{price_count2}")
            print(f"TYPE_COUNT:{type_count2}")
            print(f"📊 สรุปผลลัพธ์: Price={price_count2}, Type={type_count2}")
        except Exception as e:
            print(f"❌ Error นับ records: {e}")
            logger.error(f"Error counting records: {e}")
            print("PRICE_COUNT:0")
            print("TYPE_COUNT:0")
        
        print("SUCCESS:")
        print("🎉 ประมวลผลเสร็จสิ้นสมบูรณ์!")
        sys.exit(0)
    else:
        # Web server mode
        print("🚀 Starting Format Tostem Web Server...")
        print("📁 Upload folder:", UPLOAD_FOLDER)
        print("📁 Output folder:", OUTPUT_FOLDER)
        print("🌐 Server will be available at: http://localhost:5000")
        print("📱 You can also access from other devices at: http://[your-ip]:5000")
        print("⚠️  Press Ctrl+C to stop the server")
        print()
        
        # Install required packages if not available
        try:
            import flask
            import pandas
            import openpyxl
        except ImportError as e:
            print(f"❌ Missing required package: {e}")
            print("💡 Please install required packages:")
            print("   pip install flask pandas openpyxl")
            sys.exit(1)
        
        app.run(debug=True, host='0.0.0.0', port=5000)