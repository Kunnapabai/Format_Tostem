#!/usr/bin/env python3
"""
Excel Color Extractor - Streamlit Web Application
Simple web interface using Streamlit
"""

import os
import re
import math
import tempfile
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
import io

# Page config
st.set_page_config(
    page_title="Excel Color Extractor",
    page_icon="🎨",
    layout="wide",
    initial_sidebar_state="expanded"
)

class ColorExtractor:
    def __init__(self):
        pass
        
    def to_number(self, val):
        """Convert value to number, removing commas"""
        try:
            if val is None:
                return None
            
            str_val = str(val).strip()
            # Remove comma, space, and special characters
            clean_val = re.sub(r'[,\s]', '', str_val)
            clean_val = re.sub(r'[^\d.-]', '', clean_val)
            
            if not clean_val or clean_val in ['', '-', '.']:
                return None
                
            f = float(clean_val)
            if math.isnan(f):
                return None
            return int(f) if f.is_integer() else f
        except:
            return None

    def normalize_rgb(self, fill):
        """Convert ARGB color to RGB hex format"""
        if not fill:
            return "FFFFFF"
        
        color_found = ""
        
        # Check fgColor
        if hasattr(fill, 'fgColor') and fill.fgColor:
            if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                color_str = str(fill.fgColor.rgb).upper()
                if color_str == "00000000":
                    return "FFFFFF"
                elif len(color_str) == 8:
                    color_found = color_str[2:]
                elif len(color_str) == 6:
                    color_found = color_str
        
        # Check bgColor
        if not color_found and hasattr(fill, 'bgColor') and fill.bgColor:
            if hasattr(fill.bgColor, 'rgb') and fill.bgColor.rgb:
                color_str = str(fill.bgColor.rgb).upper()
                if color_str == "00000000":
                    return "FFFFFF"
                elif len(color_str) == 8:
                    color_found = color_str[2:]
                elif len(color_str) == 6:
                    color_found = color_str
        
        return color_found if color_found else "FFFFFF"

    def find_thickness_matrix(self, ws, raw, thickness_mm):
        """Find matrix with specific thickness label"""
        thickness_patterns = [
            rf"Thk\.{thickness_mm}\s*mm",
            rf"{thickness_mm}\s*mm",
            rf"Thickness\s*{thickness_mm}",
            rf"หนา\s*{thickness_mm}"
        ]
        
        thickness_row = thickness_col = None
        for r in range(raw.shape[0]):
            for c in range(raw.shape[1]):
                cell_val = str(raw.iat[r, c]).strip() if raw.iat[r, c] is not None else ""
                for pattern in thickness_patterns:
                    if re.search(pattern, cell_val, re.IGNORECASE):
                        thickness_row, thickness_col = r, c
                        break
                if thickness_row is not None:
                    break
            if thickness_row is not None:
                break
        
        if thickness_row is None:
            return None, None
        
        # Search for h/w header
        search_range = 15
        for r in range(max(0, thickness_row - search_range), min(raw.shape[0], thickness_row + search_range + 1)):
            for c in range(max(0, thickness_col - search_range), min(raw.shape[1], thickness_col + search_range + 1)):
                cell_val = str(raw.iat[r, c]).strip() if raw.iat[r, c] is not None else ""
                if re.search(r"\bh\s*/\s*w\b", cell_val, re.IGNORECASE):
                    return r, c
        
        return None, None

    def read_color_matrix(self, ws, raw, hr, hc, widths, heights):
        """Read colors from matrix"""
        color_map = {}
        
        for i_h, h in enumerate(heights):
            for i_w, w in enumerate(widths):
                try:
                    excel_row = hr + 2 + i_h
                    excel_col = hc + 2 + i_w
                    
                    cell = ws.cell(row=excel_row, column=excel_col)
                    color = self.normalize_rgb(cell.fill)
                    color_map[(h, w)] = color
                except Exception:
                    continue
        
        return color_map

    def process_file(self, uploaded_file):
        """Process the uploaded Excel file"""
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                temp_path = tmp_file.name
            
            base_name = os.path.splitext(uploaded_file.name)[0]
            
            xls = pd.ExcelFile(temp_path, engine="openpyxl")
            wb = load_workbook(temp_path, data_only=True)
            
            price_rows = []
            type_rows = []
            price_id = 1
            type_id = 1
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            sheets = [s for s in xls.sheet_names if s.strip().lower() != "สารบัญ"]
            total_sheets = len(sheets)
            
            for idx, sheet in enumerate(sheets):
                status_text.text(f"กำลังประมวลผล Sheet: {sheet} ({idx+1}/{total_sheets})")
                progress_bar.progress((idx + 1) / total_sheets)
                
                raw = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
                ws = wb[sheet]
                
                # Find Glass_QTY and Description
                sheet_glass_qty = 1
                sheet_description = ""
                
                for r in range(raw.shape[0]):
                    for c in range(raw.shape[1] - 1):
                        if raw.iat[r, c] is None:
                            continue
                        cell = str(raw.iat[r, c]).strip()
                        low = cell.lower()
                        
                        if low in ("glass_qty", "glass qty"):
                            next_cell = raw.iat[r, c + 1]
                            qty = self.to_number(next_cell)
                            if qty is not None:
                                sheet_glass_qty = qty
                            
                        elif low == "description":
                            desc = raw.iat[r, c + 1]
                            if desc is not None:
                                sheet_description = str(desc).strip()
                
                # Find h/w header
                locs = []
                for r in range(raw.shape[0]):
                    for c in range(raw.shape[1]):
                        if raw.iat[r, c] is None:
                            continue
                        if isinstance(raw.iat[r, c], str):
                            if re.search(r"\bh\s*/\s*w\b", raw.iat[r, c], re.IGNORECASE):
                                locs.append((r, c))
                
                if not locs:
                    st.warning(f"ไม่พบ h/w ใน sheet '{sheet}'")
                    continue
                
                hr, hc = locs[0]
                
                # Read widths and heights
                widths = []
                for c in range(hc + 1, raw.shape[1]):
                    v = self.to_number(raw.iat[hr, c])
                    if v is None:
                        break
                    widths.append(v)
                
                heights = []
                for r in range(hr + 1, raw.shape[0]):
                    h_val = self.to_number(raw.iat[r, hc])
                    if h_val is None:
                        break
                    heights.append(h_val)
                
                if not widths or not heights:
                    st.warning(f"ไม่พบข้อมูล Width/Height ใน sheet '{sheet}'")
                    continue
                
                # Find thickness matrices
                color_5mm = {}
                color_6mm = {}
                color_8mm = {}
                
                for thickness in [5, 6, 8]:
                    hr_thick, hc_thick = self.find_thickness_matrix(ws, raw, thickness)
                    if hr_thick is not None:
                        widths_thick = []
                        for c in range(hc_thick + 1, raw.shape[1]):
                            v = self.to_number(raw.iat[hr_thick, c])
                            if v is None:
                                break
                            widths_thick.append(v)
                        
                        heights_thick = []
                        for r in range(hr_thick + 1, raw.shape[0]):
                            h_val = self.to_number(raw.iat[r, hc_thick])
                            if h_val is None:
                                break
                            heights_thick.append(h_val)
                        
                        if widths_thick and heights_thick:
                            colors = self.read_color_matrix(ws, raw, hr_thick, hc_thick, widths_thick, heights_thick)
                            if thickness == 5:
                                color_5mm = colors
                            elif thickness == 6:
                                color_6mm = colors
                            elif thickness == 8:
                                color_8mm = colors
                
                # Create Type record
                type_rows.append({
                    "ID": type_id,
                    "Serie": base_name,
                    "Type": sheet.strip(),
                    "Description": sheet_description,
                    "width_min": min(widths),
                    "width_max": max(widths),
                    "height_min": min(heights),
                    "height_max": max(heights),
                })
                type_id += 1
                
                # Create Price records
                for i_h, h in enumerate(heights):
                    for i_w, w in enumerate(widths):
                        raw_price = raw.iat[hr + 1 + i_h, hc + 1 + i_w]
                        p = self.to_number(raw_price)
                        if p is None:
                            continue
                        
                        color_5 = color_5mm.get((h, w), "FFFFFF")
                        color_6 = color_6mm.get((h, w), "FFFFFF")
                        color_8 = color_8mm.get((h, w), "FFFFFF")
                        
                        price_rows.append({
                            "ID": price_id,
                            "Serie": base_name,
                            "Type": sheet.strip(),
                            "Width": w,
                            "Height": h,
                            "Price": p,
                            "Glass_QTY": sheet_glass_qty,
                            "5mm_Color": color_5,
                            "6mm_Color": color_6,
                            "8mm_Color": color_8
                        })
                        price_id += 1
            
            # Cleanup
            os.unlink(temp_path)
            progress_bar.progress(1.0)
            status_text.text("ประมวลผลเสร็จสิ้น!")
            
            return pd.DataFrame(price_rows), pd.DataFrame(type_rows)
            
        except Exception as e:
            st.error(f"เกิดข้อผิดพลาด: {str(e)}")
            return None, None

def main():
    """Main Streamlit application"""
    
    # Custom CSS
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #007bff;
        margin: 0.5rem 0;
    }
    .color-box {
        width: 30px;
        height: 30px;
        border-radius: 4px;
        display: inline-block;
        margin: 2px;
        border: 1px solid #ccc;
    }
    .success-box {
        background: linear-gradient(45deg, #d4edda, #c3e6cb);
        padding: 1rem;
        border-radius: 8px;
        border: 1px solid #c3e6cb;
        color: #155724;
        margin: 1rem 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>🎨 Excel Color Extractor</h1>
        <p>สกัดข้อมูลสีจากไฟล์ Excel และสร้างไฟล์ Price/Type อัตโนมัติ</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("### 📋 คำแนะนำการใช้งาน")
        st.markdown("""
        **โครงสร้างไฟล์ที่ต้องการ:**
        - 📊 เมทริกส์หลักที่มี `h/w` เป็นหัวตาราง
        - 🔢 ข้อมูล Width (แนวนอน) และ Height (แนวตั้ง)
        - 💰 ราคาในแต่ละเซลล์
        - 🎨 สีพื้นหลังในเซลล์
        
        **เมทริกส์ความหนา:**
        - `Thk.5 mm` สำหรับความหนา 5mm
        - `Thk.6 mm` สำหรับความหนา 6mm  
        - `Thk.8 mm` สำหรับความหนา 8mm
        
        **ข้อมูลเพิ่มเติม:**
        - `Glass_QTY` - จำนวนแก้ว
        - `Description` - คำอธิบาย
        """)
        
        st.markdown("### 📈 สถิติการใช้งาน")
        if 'processed_files' not in st.session_state:
            st.session_state.processed_files = 0
        if 'total_records' not in st.session_state:
            st.session_state.total_records = 0
            
        col1, col2 = st.columns(2)
        with col1:
            st.metric("ไฟล์ที่ประมวลผล", st.session_state.processed_files)
        with col2:
            st.metric("รายการทั้งหมด", st.session_state.total_records)
    
    # Main content
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("### 📁 อัพโหลดไฟล์")
        uploaded_file = st.file_uploader(
            "เลือกไฟล์ Excel (.xlsx)",
            type=['xlsx'],
            help="อัพโหลดไฟล์ Excel ที่มีเมทริกส์ข้อมูลและสีพื้นหลัง"
        )
        
        if uploaded_file is not None:
            # File info
            file_details = {
                "ชื่อไฟล์": uploaded_file.name,
                "ขนาดไฟล์": f"{uploaded_file.size:,} bytes",
                "ประเภท": uploaded_file.type
            }
            
            st.markdown("#### 📄 ข้อมูลไฟล์")
            for key, value in file_details.items():
                st.markdown(f"**{key}:** {value}")
            
            # Processing options
            st.markdown("#### ⚙️ ตัวเลือกการประมวลผล")
            
            col_opt1, col_opt2 = st.columns(2)
            with col_opt1:
                show_preview = st.checkbox("แสดงตัวอย่างข้อมูล", value=True)
                show_debug = st.checkbox("แสดงข้อมูล Debug", value=False)
            with col_opt2:
                show_colors = st.checkbox("แสดงรหัสสี", value=True)
                auto_download = st.checkbox("ดาวน์โหลดอัตโนมัติ", value=False)
            
            # Process button
            if st.button("🚀 เริ่มประมวลผล", type="primary", use_container_width=True):
                with st.spinner("กำลังประมวลผล..."):
                    extractor = ColorExtractor()
                    start_time = datetime.now()
                    
                    if show_debug:
                        st.markdown("#### 🐛 Debug Information")
                        debug_container = st.container()
                    
                    price_df, type_df = extractor.process_file(uploaded_file)
                    
                    end_time = datetime.now()
                    processing_time = (end_time - start_time).total_seconds()
                    
                    if price_df is not None and type_df is not None:
                        st.session_state.price_df = price_df
                        st.session_state.type_df = type_df
                        st.session_state.processing_time = processing_time
                        st.session_state.processed_files += 1
                        st.session_state.total_records += len(price_df)
                        
                        st.markdown(f"""
                        <div class="success-box">
                            <h4>✅ ประมวลผลสำเร็จ!</h4>
                            <p><strong>⏱️ เวลาประมวลผล:</strong> {processing_time:.2f} วินาที</p>
                            <p><strong>📊 รายการที่สร้าง:</strong> {len(price_df)} รายการ</p>
                        </div>
                        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("### 📊 ผลลัพธ์")
        
        if 'price_df' in st.session_state and 'type_df' in st.session_state:
            price_df = st.session_state.price_df
            type_df = st.session_state.type_df
            
            # Summary metrics
            st.markdown("#### 📈 สรุปผลลัพธ์")
            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                st.metric("📊 Price Records", len(price_df))
            with col_m2:
                st.metric("📋 Type Records", len(type_df))
            with col_m3:
                st.metric("⏱️ เวลา (วิ)", f"{st.session_state.processing_time:.2f}")
            
            # Quick stats
            if not price_df.empty:
                st.markdown("#### 📋 สถิติเร็ว")
                col_s1, col_s2 = st.columns(2)
                with col_s1:
                    st.write(f"**ราคาเฉลี่ย:** {price_df['Price'].mean():.0f}")
                    st.write(f"**ราคาสูงสุด:** {price_df['Price'].max():.0f}")
                with col_s2:
                    st.write(f"**ราคาต่ำสุด:** {price_df['Price'].min():.0f}")
                    st.write(f"**Serie:** {price_df['Serie'].iloc[0] if len(price_df) > 0 else 'N/A'}")
            
            # Download section
            st.markdown("#### 💾 ดาวน์โหลดไฟล์")
            
            col_dl1, col_dl2 = st.columns(2)
            
            with col_dl1:
                # Price file download
                price_buffer = io.BytesIO()
                price_df.to_excel(price_buffer, index=False, engine='openpyxl')
                price_buffer.seek(0)
                
                st.download_button(
                    label="📊 ดาวน์โหลด Price.xlsx",
                    data=price_buffer,
                    file_name="Price.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col_dl2:
                # Type file download
                type_buffer = io.BytesIO()
                type_df.to_excel(type_buffer, index=False, engine='openpyxl')
                type_buffer.seek(0)
                
                st.download_button(
                    label="📋 ดาวน์โหลด Type.xlsx",
                    data=type_buffer,
                    file_name="Type.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            # Preview data
            if show_preview:
                st.markdown("#### 👀 ตัวอย่างข้อมูล")
                
                tab1, tab2 = st.tabs(["📊 Price Data", "📋 Type Data"])
                
                with tab1:
                    st.dataframe(
                        price_df.head(20), 
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    if show_colors and any(col.endswith('_Color') for col in price_df.columns):
                        st.markdown("#### 🎨 แสดงสี")
                        color_cols = [col for col in price_df.columns if col.endswith('_Color')]
                        
                        for col in color_cols:
                            unique_colors = price_df[col].unique()
                            non_white_colors = [c for c in unique_colors if c and c not in ['FFFFFF', '']]
                            
                            if non_white_colors:
                                st.markdown(f"**{col}:**")
                                color_html = ""
                                for color in non_white_colors[:12]:  # Show max 12 colors
                                    color_html += f'<div class="color-box" style="background-color: #{color};" title="#{color}"></div>'
                                
                                st.markdown(color_html, unsafe_allow_html=True)
                                
                                # Show color codes
                                with st.expander(f"รหัสสี {col}"):
                                    color_df = pd.DataFrame({
                                        'Color Code': non_white_colors,
                                        'Count': [len(price_df[price_df[col] == c]) for c in non_white_colors]
                                    })
                                    st.dataframe(color_df, use_container_width=True)
                
                with tab2:
                    st.dataframe(
                        type_df, 
                        use_container_width=True,
                        hide_index=True
                    )
        else:
            st.info("📤 อัพโหลดไฟล์และกดประมวลผลเพื่อดูผลลัพธ์")
            
            # Show example
            st.markdown("#### 📋 ตัวอย่างโครงสร้างไฟล์")
            example_data = {
                'h/w': ['h/w', 800, 900, 1000],
                '100': [100, 1500, 1700, 1900],
                '120': [120, 1800, 2000, 2200],
                '150': [150, 2100, 2300, 2500]
            }
            example_df = pd.DataFrame(example_data)
            st.dataframe(example_df, use_container_width=True, hide_index=True)
            st.caption("ตัวอย่าง: เมทริกส์ที่มี h/w เป็นหัวตาราง พร้อมสีพื้นหลังในเซลล์")
    
    # Footer
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center; color: #666;'>
            <p>🎨 <strong>Excel Color Extractor v1.0.0</strong></p>
            <p>สร้างด้วย Streamlit • รองรับไฟล์ .xlsx • ลบเครื่องหมายจุลภาคในราคาอัตโนมัติ</p>
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
    